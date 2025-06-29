#!/usr/bin/env python3
"""
Multi-Sheet Excel Uploader for Multi-Tenant SQL Agent

Handles uploading Excel files with multiple sheets, converting each sheet
to a separate table in the user's dedicated PostgreSQL database.
"""

import os
import pandas as pd
from sqlalchemy import create_engine, inspect, text
from sqlalchemy.engine import Engine
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
import re
from openpyxl import load_workbook

# get_database_connection removed - using schema-per-tenant connections
from schema_user_service import SchemaUserSession as UserSession


class MultiSheetExcelUploader:
    """Handles uploading Excel files with multiple sheets to user databases."""
    
    def __init__(self):
        """Initialize the uploader."""
        self.supported_extensions = {'.xlsx', '.xls', '.csv'}
    
    def clean_name(self, name: str) -> str:
        """
        Clean names to be PostgreSQL-compatible.
        
        Args:
            name: Original name
            
        Returns:
            Cleaned name
        """
        # Convert to lowercase
        cleaned = str(name).lower()
        
        # Replace spaces and special characters with underscores
        cleaned = re.sub(r'[^a-z0-9_]', '_', cleaned)
        
        # Remove multiple consecutive underscores
        cleaned = re.sub(r'_+', '_', cleaned)
        
        # Remove leading/trailing underscores
        cleaned = cleaned.strip('_')
        
        # Ensure it doesn't start with a number
        if cleaned and cleaned[0].isdigit():
            cleaned = f'tbl_{cleaned}'
        
        # Handle empty names
        if not cleaned:
            cleaned = 'unnamed'
        
        # Ensure maximum length for PostgreSQL
        if len(cleaned) > 63:
            cleaned = cleaned[:60] + '_tr'  # tr = truncated
        
        return cleaned
    
    def generate_table_name(self, filename: str, sheet_name: str = None) -> str:
        """
        Generate a unique table name from filename and sheet name.
        
        Args:
            filename: Original filename
            sheet_name: Sheet name (for Excel files)
            
        Returns:
            Cleaned table name
        """
        # Get filename without extension
        file_base = Path(filename).stem
        file_clean = self.clean_name(file_base)
        
        if sheet_name:
            sheet_clean = self.clean_name(sheet_name)
            # Combine filename and sheet name
            table_name = f"{file_clean}_{sheet_clean}"
        else:
            table_name = file_clean
        
        # Ensure table name is valid and unique
        if not table_name:
            table_name = 'uploaded_table'
        
        return table_name
    
    def get_excel_sheets(self, file_path: str) -> List[str]:
        """
        Get list of sheet names from an Excel file.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            List of sheet names
        """
        try:
            if file_path.endswith('.csv'):
                return ['default']  # CSV files have only one "sheet"
            
            # Use openpyxl to get sheet names
            workbook = load_workbook(file_path, read_only=True)
            sheet_names = workbook.sheetnames
            workbook.close()
            
            return sheet_names
            
        except Exception as e:
            print(f"Error reading Excel sheets from {file_path}: {e}")
            return ['Sheet1']  # Default fallback
    
    def read_sheet_data(self, file_path: str, sheet_name: str = None) -> pd.DataFrame:
        """
        Read data from a specific sheet or CSV file.
        
        Args:
            file_path: Path to the file
            sheet_name: Sheet name (None for CSV)
            
        Returns:
            DataFrame with the sheet contents
        """
        file_ext = Path(file_path).suffix.lower()
        
        try:
            if file_ext == '.csv':
                # Try different encodings for CSV files
                for encoding in ['utf-8', 'latin-1', 'cp1252']:
                    try:
                        df = pd.read_csv(file_path, encoding=encoding)
                        break
                    except UnicodeDecodeError:
                        continue
                else:
                    raise Exception("Could not decode CSV file with any supported encoding")
            
            elif file_ext in ['.xlsx', '.xls']:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            
            else:
                raise ValueError(f"Unsupported file format: {file_ext}")
            
            # Clean column names
            df.columns = [self.clean_name(col) for col in df.columns]
            
            # Handle duplicate column names
            seen_columns = set()
            new_columns = []
            for col in df.columns:
                if col in seen_columns:
                    counter = 1
                    new_col = f"{col}_{counter}"
                    while new_col in seen_columns:
                        counter += 1
                        new_col = f"{col}_{counter}"
                    new_columns.append(new_col)
                    seen_columns.add(new_col)
                else:
                    new_columns.append(col)
                    seen_columns.add(col)
            
            df.columns = new_columns
            
            # Remove completely empty rows
            df = df.dropna(how='all')
            
            return df
        
        except Exception as e:
            raise Exception(f"Error reading sheet '{sheet_name}' from {file_path}: {str(e)}")
    
    def upload_file_with_sheets(self, file_path: str, user_session: UserSession) -> Dict[str, Any]:
        """
        Upload a file with all its sheets to the user's database.
        
        Args:
            file_path: Path to the file to upload
            user_session: User session with database info
            
        Returns:
            Dictionary with upload results
        """
        try:
            filename = Path(file_path).name
            file_ext = Path(file_path).suffix.lower()
            
            if file_ext not in self.supported_extensions:
                return {
                    'success': False,
                    'error': f'Unsupported file format: {file_ext}'
                }
            
            # Get schema-per-tenant database connection
            from sqlalchemy import create_engine
            engine = create_engine(user_session.db_uri)
            
            # Get all sheets in the file
            sheet_names = self.get_excel_sheets(file_path)
            
            results = {
                'success': True,
                'message': '',
                'tables_created': [],
                'total_rows': 0,
                'sheets_processed': 0,
                'errors': []
            }
            
            for sheet_name in sheet_names:
                try:
                    # Read sheet data
                    df = self.read_sheet_data(file_path, sheet_name if file_ext != '.csv' else None)
                    
                    if df.empty:
                        results['errors'].append(f"Sheet '{sheet_name}' is empty")
                        continue
                    
                    # Generate table name
                    table_name = self.generate_table_name(filename, sheet_name if file_ext != '.csv' else None)
                    
                    # Ensure table name is unique in this database
                    original_table_name = table_name
                    counter = 1
                    while self.table_exists(engine, table_name):
                        table_name = f"{original_table_name}_{counter}"
                        counter += 1
                    
                    # Upload to database using simplified approach for compatibility
                    try:
                        # Simple approach: Use connection context manager
                        conn = engine.connect()
                        try:
                            df.to_sql(
                                table_name,
                                con=conn,
                                if_exists='replace',
                                index=False,
                                method='multi'
                            )
                            print(f"✅ Successfully uploaded {table_name} with {len(df)} rows")
                        finally:
                            conn.close()
                    except Exception as e:
                        # Use manual approach with simplified type handling (don't log this as error)
                        try:
                            self._simple_manual_upload(df, table_name, engine)
                            # Manual upload succeeded, continue normally
                        except Exception as manual_error:
                            # Only raise if manual approach also fails
                            raise manual_error
                    
                    # Record the uploaded table
                    user_session.add_uploaded_table(
                        table_name=table_name,
                        source_file=filename,
                        sheet_name=sheet_name if file_ext != '.csv' else None
                    )
                    
                    results['tables_created'].append(table_name)
                    results['total_rows'] += len(df)
                    results['sheets_processed'] += 1
                    
                except Exception as e:
                    error_msg = f"Sheet '{sheet_name}': {str(e)}"
                    results['errors'].append(error_msg)
                    print(f"Error uploading sheet '{sheet_name}': {e}")
            
            # Update overall success status and message
            if results['sheets_processed'] == 0:
                results['success'] = False
                results['message'] = f"Failed to upload any sheets from {filename}"
            elif results['errors']:
                results['message'] = f"Partially successful: {results['sheets_processed']} sheets uploaded from {filename}, {len(results['errors'])} failed"
            else:
                results['message'] = f"Successfully uploaded all {results['sheets_processed']} sheets from {filename} with {results['total_rows']} total rows"
            
            return results
        
        except Exception as e:
            return {
                'success': False,
                'error': f"Failed to process file {filename}: {str(e)}"
            }
    
    def table_exists(self, engine: Engine, table_name: str) -> bool:
        """
        Check if a table exists in the database.
        
        Args:
            engine: SQLAlchemy engine
            table_name: Name of the table to check
            
        Returns:
            True if table exists, False otherwise
        """
        try:
            inspector = inspect(engine)
            return table_name in inspector.get_table_names()
        except Exception:
            return False
    
    def upload_multiple_files(self, file_paths: List[str], user_session: UserSession) -> Dict[str, Any]:
        """
        Upload multiple files to the user's database.
        
        Args:
            file_paths: List of file paths to upload
            user_session: User session with database info
            
        Returns:
            Dictionary with upload results
        """
        overall_results = {
            'success': True,
            'message': '',
            'tables_created': [],
            'total_rows': 0,
            'files_processed': 0,
            'sheets_processed': 0,
            'errors': []
        }
        
        for file_path in file_paths:
            try:
                result = self.upload_file_with_sheets(file_path, user_session)
                
                if result['success']:
                    overall_results['tables_created'].extend(result['tables_created'])
                    overall_results['total_rows'] += result['total_rows']
                    overall_results['sheets_processed'] += result['sheets_processed']
                    overall_results['files_processed'] += 1
                else:
                    filename = Path(file_path).name
                    overall_results['errors'].append(f"{filename}: {result.get('error', 'Unknown error')}")
                
                # Add any sheet-level errors
                if 'errors' in result:
                    overall_results['errors'].extend([f"{Path(file_path).name} - {err}" for err in result['errors']])
                    
            except Exception as e:
                filename = Path(file_path).name
                overall_results['errors'].append(f"{filename}: {str(e)}")
        
        # Update overall success status
        if overall_results['errors']:
            if overall_results['files_processed'] == 0:
                overall_results['success'] = False
                overall_results['message'] = f"Failed to upload any files. Errors: {'; '.join(overall_results['errors'][:3])}"
            else:
                overall_results['message'] = f"Partially successful: {overall_results['files_processed']} files, {overall_results['sheets_processed']} sheets uploaded, {len(overall_results['errors'])} errors"
        else:
            overall_results['message'] = f"Successfully uploaded {overall_results['files_processed']} files with {overall_results['sheets_processed']} sheets and {overall_results['total_rows']} total rows"
        
        return overall_results
    
    def get_user_tables(self, user_session: UserSession) -> List[Dict[str, Any]]:
        """
        Get list of tables in the user's database.
        
        Args:
            user_session: User session with database info
            
        Returns:
            List of table information
        """
        try:
            from sqlalchemy import create_engine
            engine = create_engine(user_session.db_uri)
            inspector = inspect(engine)
            
            tables = []
            for table_name in inspector.get_table_names():
                # Get table info from session if available
                table_info = next(
                    (t for t in user_session.uploaded_tables if t['table_name'] == table_name),
                    None
                )
                
                if table_info:
                    tables.append(table_info)
                else:
                    # Table exists but not in session history (maybe from previous session)
                    tables.append({
                        'table_name': table_name,
                        'source_file': 'Unknown',
                        'sheet_name': None,
                        'uploaded_at': 'Unknown'
                    })
            
            return tables
            
        except Exception as e:
            print(f"Error getting user tables: {e}")
            return []
    
    def _simple_manual_upload(self, df, table_name, engine):
        """Manual SQL upload approach with simplified type handling."""
        try:
            with engine.connect() as connection:
                trans = connection.begin()
                try:
                    # Drop table if exists
                    connection.execute(text(f"DROP TABLE IF EXISTS {table_name}"))
                    
                    # Create table with simplified column types (all TEXT for compatibility)
                    column_definitions = []
                    for col in df.columns:
                        column_definitions.append(f'"{col}" TEXT')
                    
                    create_sql = f"CREATE TABLE {table_name} ({', '.join(column_definitions)})"
                    connection.execute(text(create_sql))
                    
                    # Insert data
                    if not df.empty:
                        columns = ', '.join([f'"{col}"' for col in df.columns])
                        placeholders = ', '.join([f':col_{i}' for i in range(len(df.columns))])
                        insert_sql = f"INSERT INTO {table_name} ({columns}) VALUES ({placeholders})"
                        
                        for index, row in df.iterrows():
                            # Convert row values to strings, handling mixed types
                            values = {}
                            for i, val in enumerate(row):
                                if pd.isna(val):
                                    values[f'col_{i}'] = None
                                else:
                                    values[f'col_{i}'] = str(val)
                            connection.execute(text(insert_sql), values)
                    
                    trans.commit()
                    print(f"✅ Manual upload successful for table {table_name} with {len(df)} rows")
                    
                except Exception as e:
                    trans.rollback()
                    raise e
                    
        except Exception as e:
            print(f"❌ Error in manual upload: {e}")
            raise e


def main():
    """Example usage of the multi-sheet uploader."""
    from user_service import get_user_service
    
    # Create a test user session
    user_service = get_user_service()
    session = user_service.create_user_session()
    
    uploader = MultiSheetExcelUploader()
    
    # Example file upload
    file_path = "example.xlsx"  # Replace with actual file path
    
    if os.path.exists(file_path):
        result = uploader.upload_file_with_sheets(file_path, session)
        print(f"Upload result: {result}")
        
        # List tables
        tables = uploader.get_user_tables(session)
        print(f"User tables: {tables}")
    else:
        print(f"File {file_path} not found")


if __name__ == "__main__":
    main()
